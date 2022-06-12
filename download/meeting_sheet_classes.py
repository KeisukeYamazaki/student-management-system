from decimal import Decimal
from decimal import ROUND_HALF_UP
from logging import getLogger

from openpyxl.styles import PatternFill
from openpyxl.styles import Font

from student.exam_record import ExamRecord
from student.models import JuniorHighSchool
from util import get_school_year
from util import str_for_int_check
from util import str_for_float_check


logger = getLogger(__name__)


class MeetingSheetCommon:
    def __init__(self, student, ws):
        self.exam_record = ExamRecord(student)
        self.student = student
        self.term_system = JuniorHighSchool.objects.get(name=student.local_school).term
        self.ws = ws
        self.fill = PatternFill(patternType='solid', fgColor='d3d3d3')
        self.font = Font(name='游明朝', size=10, color='FF0000')
        # ループするときに A = 1 となるように、リストの最初は空文字にする
        self.col_list = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
                         'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']
        self.school_dict_keys = ['first', 'second', 'third', 'fourth', 'fifth']
        self.SCHOOL_RECORD_START_COL = self.P_EXAM_START_COL = 6
        self.SCHOOL_RECORD_END_COL = 16 + 1
        self.P_EXAM_END_COL = 13 + 1

        self.R_EXAM_START_COL = 19
        self.R_EXAM_END_COL = 29 + 1

        self.PRI_SCHOOL_ROW = 28
        self.PRI_SCHOOL_ROW2 = self.PRI_SCHOOL_ROW + 20

        self.THREE_FIVE_ROW = 14

        if self.term_system == '３学期制':
            self.SCHOOL_RECORD_ROW1 = 4
            self.SCHOOL_RECORD_ROW2 = 7
            self.SCHOOL_RECORD_ROW3 = 10

            self.R_EXAM_ROW1 = 4
            self.R_EXAM_ROW2 = 9
            self.R_EXAM_ROW3 = 14

            self.P_EXAM_ROW = 15  # その学年になってからの初めての模試(７月)を開始行とする(中２,中３)

            self.TERMS = ['１学期', '２学期', '３学期']
            self.R_EXAMS = ['１学期中間', '１学期期末', '２学期中間', '２学期期末', '学年末']

        elif self.term_system == '２学期制':
            self.SCHOOL_RECORD_ROW1 = 4
            self.SCHOOL_RECORD_ROW2 = 6
            self.SCHOOL_RECORD_ROW3 = 8

            self.R_EXAM_ROW1 = 4
            self.R_EXAM_ROW2 = 8
            self.R_EXAM_ROW3 = 12

            self.P_EXAM_ROW = 13  # その学年になってからの初めての模試(７月)を開始行とする(中２,中３)

            self.TERMS = ['前期', '後期']
            self.R_EXAMS = ['前期中間', '前期期末', '後期中間', '学年末']

        if self.student.grade == '中１':
            self.PRI_SCHOOL_ROW2 += 4
            self.THREE_FIVE_ROW = 12

        self.ws.cell(row=1, column=1, value=f'{str(get_school_year())}年度')
        self.ws.cell(row=2, column=23, value=f'{student.last_name} {student.first_name}')
        self.ws.cell(row=4, column=2, value=self.exam_record.one_three_five)
        self.ws.cell(row=6, column=2, value=self.exam_record.four_five)
        self.ws.cell(row=8, column=2, value=self.exam_record.five_zero)
        self.ws.cell(row=10, column=2, value=self.exam_record.two_five)

    def set_A_value(self):
        A_value = self.exam_record.get_A_score()
        if not A_value == '':
            self.ws.cell(row=12, column=2, value=A_value.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP))
        else:
            self.ws.cell(row=12, column=2, value='')

    def set_school_record_one_row(self, school_record, row):
        """set_school_record_one_row

        １行分の成績を書き込む

        Args:
            school_record (SchoolRecord): SchoolRecordインスタンス
            row (int): 行番号
        """
        record_content_list = school_record.to_list()
        SUBJECT_START = 4
        col = self.SCHOOL_RECORD_START_COL
        for i in range(SUBJECT_START, len(record_content_list)):
            if str_for_int_check(record_content_list[i]):
                self.ws.cell(row=row, column=col, value=int(record_content_list[i]))
                if int(record_content_list[i]) == 5:
                    self.ws[f'{self.col_list[col]}{row}'].font = self.font
            else:
                self.ws.cell(row=row, column=col, value=record_content_list[i])
            col += 1

    def set_regular_exam_one_row(self, regular_exam, row):
        """set_regular_exam_one_row

        １行分の定期試験結果を書き込む

        Args:
            regular_exam (RegularExam): RegularExamインスタンス
            row (int): 行番号
        """
        r_exam_list = regular_exam.to_list()
        SUBJECT_START = 4
        col = self.R_EXAM_START_COL
        for i in range(SUBJECT_START, len(r_exam_list)):
            if str_for_int_check(r_exam_list[i]):
                self.ws.cell(row=row, column=col, value=int(r_exam_list[i]))
                if 90 <= int(r_exam_list[i]) <= 100:
                    self.ws[f'{self.col_list[col]}{row}'].font = self.font
            else:
                self.ws.cell(row=row, column=col, value=r_exam_list[i])
            col += 1

    def set_practice_exam_one_row(self, practice_exam, exam_record, row):
        """set_practice_exam_one_row

        １行分の模試結果を書き込む

        Args:
            practice_exam (PracticeExam): PracticeExamインスタンス
            exam_record (ExamRecord): ExamRecordインスタンス
            row (int): 行番号
        """
        p_exam_list = practice_exam.to_list()
        SUBJECT_START = 4
        col = self.P_EXAM_START_COL
        afbg_dict = practice_exam.get_afbg(exam_record)
        for i in range(SUBJECT_START, len(p_exam_list)):
            if str_for_int_check(p_exam_list[i]):
                self.ws.cell(row=row, column=col, value=int(p_exam_list[i]))
                if 90 <= int(p_exam_list[i]) <= 100:
                    self.ws[f'{self.col_list[col]}{row}'].font = self.font
            else:
                self.ws.cell(row=row, column=col, value=p_exam_list[i])
            col += 1

            self.ws.cell(row=row, column=14, value=afbg_dict['three_five'])
            self.ws.cell(row=self.THREE_FIVE_ROW, column=2, value=afbg_dict['three_five'])
            self.ws.cell(row=row, column=15, value=afbg_dict['four_four'])
            self.ws.cell(row=self.THREE_FIVE_ROW + 2, column=2, value=afbg_dict['four_four'])
            self.ws.cell(row=row, column=16, value=afbg_dict['five_three'])
            self.ws.cell(row=self.THREE_FIVE_ROW + 4, column=2, value=afbg_dict['five_three'])

    def fill_blank_one_row(self, row, start_col, end_col):
        """fill_blank_one_row

        空行を指定の色で埋める

        Args:
            row (int): 行番号
            start_col (int) : ループ開始列
            end_col (int) : ループ終了列
        """
        for i in range(start_col, end_col):
            self.ws[f'{self.col_list[i]}{row}'].fill = self.fill

    def set_school_record(self, record_list, write_grade, now_grade):
        """set_school_record

        当該の学年の成績を書き込む

        Args:
            record_list (list): SchoolRecordを格納したリスト
            write_grade (int) : 成績を書き込む対象の学年(1~3)
            now_grade (int) : 現在の学年(1~3)
        """
        if write_grade == 3:
            row = self.SCHOOL_RECORD_ROW3
        elif write_grade == 2:
            row = self.SCHOOL_RECORD_ROW2
        elif write_grade == 1:
            row = self.SCHOOL_RECORD_ROW1

        # 成績が１つ以上ある場合
        for j in range(len(self.TERMS)):
            term_name = self.TERMS[j]
            if len(record_list) == 0 and not write_grade == now_grade:
                # 成績がない かつ 書き込む対象の学年と現在の学年が等しくない場合、空行を埋める
                self.fill_blank_one_row(row + j, self.SCHOOL_RECORD_START_COL, self.SCHOOL_RECORD_END_COL)
            for list_num in range(len(record_list)):
                if record_list[list_num].term_name.term_name == term_name:
                    self.set_school_record_one_row(record_list[list_num], row + j)
                    break
                if list_num == len(record_list) - 1:
                    # 該当の成績がない場合は空行を埋める
                    self.fill_blank_one_row(row + j, self.SCHOOL_RECORD_START_COL, self.SCHOOL_RECORD_END_COL)

    def set_regular_exam(self, r_exam_list, write_grade, now_grade):
        """set_regular_exam

        当該の学年の成績を書き込む

        Args:
            r_exam_list (list): RegularExamを格納したリスト
            write_grade (int) : 試験結果を書き込む対象の学年(1~3)
            now_grade (int) : 現在の学年(1~3)
        """
        if write_grade == 3:
            row = self.R_EXAM_ROW3
        elif write_grade == 2:
            row = self.R_EXAM_ROW2
        elif write_grade == 1:
            row = self.R_EXAM_ROW1

        for j in range(len(self.R_EXAMS)):
            exam_name = self.R_EXAMS[j]
            if len(r_exam_list) == 0 and not write_grade == now_grade:
                # 試験結果がない かつ 書き込む対象の学年と現在の学年が等しくない場合、空行を埋める
                self.fill_blank_one_row(row + j, self.R_EXAM_START_COL, self.R_EXAM_END_COL)
            for exam_num in range(len(r_exam_list)):
                if r_exam_list[exam_num].regular_id.name == exam_name:
                    self.set_regular_exam_one_row(r_exam_list[exam_num], row + j)
                    break
                last = len(r_exam_list) - 1
                if exam_num == last:
                    # 空行は埋める
                    self.fill_blank_one_row(row + j, self.R_EXAM_START_COL, self.R_EXAM_END_COL)

    def set_practice_exam(self, this_grade_exam_list, exam_record, last_grade_exam=None):
        if self.student.grade == '中３':
            self.PRACTICE_EXAMS = ['７月', '８月', '10月', '12月', '１月']
            if last_grade_exam:
                self.set_practice_exam_one_row(last_grade_exam, exam_record, self.P_EXAM_ROW - 1)
            else:
                self.fill_blank_one_row(self.P_EXAM_ROW - 1, self.P_EXAM_START_COL, self.P_EXAM_END_COL + 3)
        elif self.student.grade == '中２':
            self.PRACTICE_EXAMS = ['７月', '８月', '10月', '12月', '３月']
            if last_grade_exam:
                self.set_practice_exam_one_row(last_grade_exam, exam_record, self.P_EXAM_ROW - 1)
            else:
                self.fill_blank_one_row(self.P_EXAM_ROW - 1, self.P_EXAM_START_COL, self.P_EXAM_END_COL + 3)
        elif self.student.grade == '中１':
            self.PRACTICE_EXAMS = ['８月', '10月', '12月', '３月']
            self.P_EXAM_ROW -= 1

        for exam_num in range(len(this_grade_exam_list)):
            # term_nameで判断して、記入する行を分岐していく
            for j in range(len(self.PRACTICE_EXAMS)):
                if this_grade_exam_list[exam_num].month.name == self.PRACTICE_EXAMS[j]:
                    self.set_practice_exam_one_row(this_grade_exam_list[exam_num], exam_record, self.P_EXAM_ROW + j)
                    break
                if j == len(self.PRACTICE_EXAMS) - 1:
                    # 空行は埋める(END_COLは、比率の部分も埋めるので +3 とする)
                    self.fill_blank_one_row(self.P_EXAM_ROW + j, self.P_EXAM_START_COL, self.P_EXAM_END_COL + 3)

    def set_private_school(self, private_school_dict):
        """set_school_record_one_row

        候補となっている私立高校データをすべて書き込む

        Args:
            private_school_dict (dict): PrivateHighSchoolを格納したdict
        """
        # 面談シート１ページ目の私立高校の部分
        school_list_num = 0
        for row in range(self.PRI_SCHOOL_ROW, self.PRI_SCHOOL_ROW + len(private_school_dict)):
            school_content_list = private_school_dict[self.school_dict_keys[school_list_num]].to_list()
            self.ws.cell(row=row, column=1, value=school_content_list[0])  # 高校名
            self.ws.cell(row=row, column=3, value=school_content_list[1])  # 学科・コース
            self.ws.cell(row=row, column=6, value=school_content_list[2])  # 基準値
            school_list_num += 1
        # 面談シート２ページ目
        school_list_num = 0
        for row2 in range(self.PRI_SCHOOL_ROW2, self.PRI_SCHOOL_ROW2 + len(private_school_dict)):
            school_content_list = private_school_dict[self.school_dict_keys[school_list_num]].to_list()
            self.ws.cell(row=row2, column=1, value=school_content_list[0])  # 高校名
            self.ws.cell(row=row2, column=3, value=school_content_list[3])  # アクセス
            school_list_num += 1


class MeetingSheet3rd2nd(MeetingSheetCommon):
    def __init__(self, ws, student):
        super().__init__(student, ws)
        self.PUB_SCHOOL_ROW = 23
        self.PUB_SCHOOL_ROW2 = self.PUB_SCHOOL_ROW + 13
        self.PUB_SCHOOL_ROW3 = self.PUB_SCHOOL_ROW + 20
        self.PUBLIC_NUM_START_COL = 5  # 公立高校の数値の入力が始まる列
        self.PUBLIC_NUM_END_COL = 17  # 公立高校の数値の入力が終わる列 + 1

    def set_public_school(self, public_school_dict, calc_score_dict_list):
        """set_school_record_one_row

        候補となっている公立高校データをすべて書き込む

        Args:
            public_school_dict (dict): PublicHighSchoolを格納したdict
            calc_score_dict_list (list): 平均内申までの差、目標点、ボーダー点のdictを格納したlist
        """
        school_list_num = 0
        # 面談シート１ページ目の公立高校の部分
        logger.info('面談シート１ページ目の公立高校の部分')
        for row1 in range(self.PUB_SCHOOL_ROW, self.PUB_SCHOOL_ROW + len(public_school_dict)):
            school_content_list = public_school_dict[self.school_dict_keys[school_list_num]].to_list()

            # 高校名が空ならとばす
            if school_content_list[0] == '':
                continue

            logger.info(f'{self.school_dict_keys[school_list_num]}:{school_content_list}')

            self.ws.cell(row=row1, column=1, value=school_content_list[0])  # 高校名
            self.ws.cell(row=row1, column=3, value=school_content_list[1])  # 比率
            school_list_num += 1
            content_list_num = 2
            for column in range(self.PUBLIC_NUM_START_COL, self.PUBLIC_NUM_END_COL):
                if str_for_float_check(school_content_list[content_list_num]):
                    self.ws.cell(row=row1, column=column, value=float(school_content_list[content_list_num]))
                else:
                    self.ws.cell(row=row1, column=column, value=school_content_list[content_list_num])
                content_list_num += 1
        # 面談シート２ページ目：「現在の内申での入試必要点数」の部分
        list_num = 0
        logger.info('面談シート２ページ目：「現在の内申での入試必要点数」')
        for row2 in range(self.PUB_SCHOOL_ROW2, self.PUB_SCHOOL_ROW2 + len(public_school_dict)):
            school_content_list = public_school_dict[self.school_dict_keys[list_num]].to_list()

            # 高校名が空ならとばす
            if school_content_list[0] == '':
                school_list_num += 1
                continue

            self.ws.cell(row=row2, column=1, value=school_content_list[0])  # 高校名
            try:
                self.ws.cell(row=row2, column=3, value=calc_score_dict_list[list_num]['ave']['school_record'])  # 平均内申までの差
                self.ws.cell(row=row2, column=8, value=calc_score_dict_list[list_num]['ave']['exam_score'])  # 目標点
                self.ws.cell(row=row2, column=12, value=calc_score_dict_list[list_num]['ave']['border'])  # ボーダー点
            except KeyError:
                pass
            list_num += 1
        # 面談シート２ページ目：「高校情報」の部分
        school_list_num = 0
        logger.info('面談シート２ページ目：「高校情報」の部分')
        for row3 in range(self.PUB_SCHOOL_ROW3, self.PUB_SCHOOL_ROW3 + len(public_school_dict)):
            school_content_list = public_school_dict[self.school_dict_keys[school_list_num]].to_list()

            # 高校名が空ならとばす
            if school_content_list[0] == '':
                school_list_num += 1
                continue

            self.ws.cell(row=row3, column=1, value=school_content_list[0])  # 高校名
            self.ws.cell(row=row3, column=3, value=school_content_list[14])  # アクセス
            school_list_num += 1


class MeetingSheet1st(MeetingSheetCommon):
    def __init__(self, ws, student):
        super().__init__(student, ws)
        self.PUB_SCHOOL_ROW = 21
        self.PUB_SCHOOL_ROW2 = self.PUB_SCHOOL_ROW + 15
        self.PUB_SCHOOL_ROW3 = self.PUB_SCHOOL_ROW + 24
        self.PRI_SCHOOL_ROW = 28
        self.PRI_SCHOOL_ROW2 = self.PRI_SCHOOL_ROW + 24

    def set_public_school(self, public_school_dict, score_dict_list, calc_score_dict_list):
        # 面談シート１ページ目の公立高校の部分
        list_num = 0
        logger.info('面談シート１ページ目の公立高校の部分')
        for row1 in range(self.PUB_SCHOOL_ROW, self.PUB_SCHOOL_ROW + len(public_school_dict)):
            school_content_list = public_school_dict[self.school_dict_keys[list_num]].to_list()
            # 高校名が空ならとばす
            if school_content_list[0] == '':
                continue
                list_num += 1

            logger.info(f'{self.school_dict_keys[list_num]}:{school_content_list}')

            try:
                raw_average = Decimal(str(score_dict_list[list_num]['ave']['school_record'])) / Decimal('3')
                average = raw_average.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
            except Exception:
                list_num += 1
                continue
            self.ws.cell(row=row1, column=1, value=school_content_list[0])  # 高校名
            self.ws.cell(row=row1, column=3, value=school_content_list[1])  # 比率
            self.ws.cell(row=row1, column=6, value='９科目合計：')

            self.ws.cell(row=row1, column=9, value=average)  # 合格者の平均内申
            self.ws.cell(row=row1, column=11, value='５科目合計：')
            self.ws.cell(row=row1, column=15, value=score_dict_list[list_num]['ave']['exam_score'])  # 合格者の学力検査平均
            list_num += 1
        # 面談シート２ページ目の公立高校１
        list_num = 0
        logger.info('面談シート２ページ目の公立高校')
        for row2 in range(self.PUB_SCHOOL_ROW2, self.PUB_SCHOOL_ROW2 + len(public_school_dict)):
            school_content_list = public_school_dict[self.school_dict_keys[list_num]].to_list()

            if school_content_list[0] == '':
                list_num += 1
                continue

            self.ws.cell(row=row2, column=1, value=school_content_list[0])  # 高校名

            try:
                self.ws.cell(row=row2, column=3, value=calc_score_dict_list[list_num]['ave']['school_record'])  # 平均内申までの差
                self.ws.cell(row=row2, column=8, value=calc_score_dict_list[list_num]['ave']['exam_score'])  # 目標点
            except KeyError:
                # 何も入力せずに終える
                pass
            list_num += 1
        # 面談シート２ページ目：「高校情報」の部分
        school_list_num = 0
        logger.info('面談シート２ページ目：「高校情報」の部分')
        for row3 in range(self.PUB_SCHOOL_ROW3, self.PUB_SCHOOL_ROW3 + len(public_school_dict)):
            school_content_list = public_school_dict[self.school_dict_keys[school_list_num]].to_list()

            if school_content_list[0] == '':
                school_list_num += 1
                continue

            self.ws.cell(row=row3, column=1, value=school_content_list[0])  # 高校名
            self.ws.cell(row=row3, column=3, value=school_content_list[14])  # アクセス
            school_list_num += 1
