import os

import openpyxl


def delete_row_with_merged_ranges(sheet, idx):
    """delete_row_with_merged_ranges

        行を書式ごと削除する

        Args:
            sheet (str): 対象のシート
            idx (int): 削除する行
    """
    begin_col, end_col = [], []
    # 削除対象行の結合セルの列情報を取得
    for n in sheet.merged_cells:
        if n.min_row == idx and n.max_row == idx:
            begin_col.append(n.min_col)
            end_col.append(n.max_col)
    # 削除対象行のセル結合を解除
    for n in range(0, len(begin_col), 1):
        sheet.unmerge_cells(
            start_row=idx,
            start_column=begin_col[n],
            end_row=idx,
            end_column=end_col[n]
        )
    sheet.delete_rows(idx)
    for mcr in sheet.merged_cells:
        if idx < mcr.min_row:
            mcr.shift(row_shift=-1)
        elif idx <= mcr.max_row:
            mcr.shrink(bottom=1)


def make_excel_name_sheet(file_name, dict_home_room_students):
    base_path = os.getcwd()

    home_rooms = list(dict_home_room_students.keys())
    home_rooms_num = len(home_rooms)

    ONE_PAGE_ROWS = 32
    ONE_CLASS_ROW = ONE_PAGE_ROWS * 2

    file_path = '/download/excel/name_sheet/meiboTemplate{}.xlsx'.format(str(home_rooms_num))
    wb = openpyxl.load_workbook(base_path + file_path)
    ws = wb.worksheets[0]

    list_end = ONE_CLASS_ROW * home_rooms_num
    start_row_list = [i for i in range(0, list_end, ONE_PAGE_ROWS * 2)]

    for i in range(home_rooms_num):
        home_room_row = start_row_list[i] + 2
        ws.cell(row=home_room_row, column=1).value = home_rooms[i]
        ws.cell(row=home_room_row + ONE_PAGE_ROWS, column=1).value = home_rooms[i]

        student_row = start_row_list[i] + 4
        for student in dict_home_room_students[home_rooms[i]]:
            student_fullname = '{} {}'.format(student.last_name, student.first_name)
            ws.cell(row=student_row, column=2).value = student_fullname
            ws.cell(row=student_row + ONE_PAGE_ROWS, column=2).value = student_fullname
            student_row += 1

    wb.save(base_path + '/' + file_name)
